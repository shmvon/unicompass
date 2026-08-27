import unittest
import os
import json
from app import app, compute_cell_colors, ANNUAL_VARIABLES
from auth import (
    authenticate,
    load_users,
    get_user,
    set_user_pincode,
    is_user_blocked,
    get_all_sectors,
    get_all_org_sector_combinations,
)
from models import (
    init_db, get_annual_results, save_annual_results,
    get_agreements, save_agreements, get_companies, save_companies,
    compute_progress, export_all, generate_export_zip,
    clear_login_attempts, record_failed_login
)

class UniCompassTestCase(unittest.TestCase):
    def setUp(self):
        app.config["TESTING"] = True
        app.config["WTF_CSRF_ENABLED"] = False
        self.client = app.test_client()
        init_db()

    def test_auth_users_csv(self):
        users = load_users()
        self.assertTrue(len(users) > 0)
        # Check system user
        u = authenticate("system@example.com", "0000")
        self.assertIsNotNone(u)
        self.assertTrue(u["is_admin"])
        self.assertIn("Commerce", u["sectors"])

        # Check BE-ACV-Finance
        u2 = authenticate("BE-ACV-Finance", "3456")
        self.assertIsNotNone(u2)
        self.assertEqual(u2["organisation"], "UNI-Belgium")
        self.assertEqual(u2["sector"], "Finance")

    def test_sector_scoped_data_storage(self):
        org = "TestOrg"
        sec1 = "Finance"
        sec2 = "Commerce"
        year = 2026

        # Save data in Finance sector
        save_annual_results(org, [{
            "variable": "bargaining_climate",
            "year": year,
            "value": "",
            "qual_value": "very favourable",
            "row_mode": "qual",
            "comment": "Good year"
        }], sector=sec1)

        # Save different data in Commerce sector
        save_annual_results(org, [{
            "variable": "bargaining_climate",
            "year": year,
            "value": "",
            "qual_value": "very unfavourable",
            "row_mode": "qual",
            "comment": "Tough year"
        }], sector=sec2)

        # Retrieve Finance
        res1 = get_annual_results(org, sector=sec1)
        self.assertEqual(res1[("bargaining_climate", year)]["qual_value"], "very favourable")
        self.assertEqual(res1[("bargaining_climate", year)]["comment"], "Good year")

        # Retrieve Commerce
        res2 = get_annual_results(org, sector=sec2)
        self.assertEqual(res2[("bargaining_climate", year)]["qual_value"], "very unfavourable")
        self.assertEqual(res2[("bargaining_climate", year)]["comment"], "Tough year")

    def test_likert_heatmap_colors_blue_to_orange(self):
        saved = {
            ("bargaining_climate", 2024): {"qual_value": "very favourable", "value": ""},
            ("bargaining_climate", 2025): {"qual_value": "neutral", "value": ""},
            ("bargaining_climate", 2026): {"qual_value": "very unfavourable", "value": ""},
            ("avg_pay_increase", 2024): {"value": "2.0", "qual_value": ""},
            ("avg_pay_increase", 2025): {"value": "4.0", "qual_value": ""},
        }
        colormap = compute_cell_colors(saved, ANNUAL_VARIABLES, [2024, 2025, 2026])
        # Very favourable -> High / Orange (249, 115, 22)
        self.assertIn("rgba(249, 115, 22", colormap[("bargaining_climate", 2024)])
        # Neutral -> Slate (148, 163, 184)
        self.assertIn("rgba(148, 163, 184", colormap[("bargaining_climate", 2025)])
        # Very unfavourable -> Low / Blue (37, 99, 235)
        self.assertIn("rgba(37, 99, 235", colormap[("bargaining_climate", 2026)])
        # Numeric min (Blue) and max (Orange)
        self.assertIn("37", colormap[("avg_pay_increase", 2024)])
        self.assertIn("249", colormap[("avg_pay_increase", 2025)])

    def test_home_page_progress_bars_and_topbar(self):
        with self.client:
            self.client.post("/login", data={"user": "system@example.com", "pincode": "0000"}, follow_redirects=True)
            res = self.client.get("/home")
            self.assertEqual(res.status_code, 200)
            html = res.get_data(as_text=True)
            self.assertIn("Input progress", html)
            # Verify no numbers like '/10' in the progress table
            self.assertNotIn("/10", html)
            # Verify topbar Logout mention with hover username
            self.assertIn("Logout", html)
            self.assertIn('title="system@example.com"', html)

    def test_admin_dashboard_org_sector_combinations(self):
        combos = get_all_org_sector_combinations()
        self.assertTrue(len(combos) > 0)
        with self.client:
            self.client.post("/login", data={"user": "admin@example.com", "pincode": "1234"}, follow_redirects=True)
            res = self.client.get("/admin")
            self.assertEqual(res.status_code, 200)
            html = res.get_data(as_text=True)
            self.assertIn("UNI-Belgium", html)
            self.assertIn("ABVV", html)
            self.assertIn("Finance", html)
            self.assertIn("Commerce", html)

    def test_exports_scoped_and_admin(self):
        # Save sample data
        save_annual_results("UNI-Belgium", [{
            "variable": "collective_bargaining_coverage",
            "year": 2026,
            "value": "85.5",
            "qual_value": "",
            "row_mode": "num",
            "comment": "",
            "sector": "Finance"
        }], sector="Finance")

        export_all()

        with self.client:
            # Regular user export view
            self.client.post("/login", data={"user": "BE-ACV-Finance", "pincode": "3456"}, follow_redirects=True)
            res = self.client.get("/export")
            self.assertEqual(res.status_code, 200)
            html = res.get_data(as_text=True)
            self.assertIn("UNI-Belgium_Finance_annual_results.csv", html)

            # Regular user cannot download other files
            res_forbid = self.client.get("/export/download/ABVV_Finance_annual_results.csv")
            self.assertEqual(res_forbid.status_code, 403)

        with self.client:
            # Admin export view (shows scoped files matching active view_org/sector, same as viewing users)
            self.client.post("/login", data={"user": "admin@example.com", "pincode": "1234"}, follow_redirects=True)
            res_admin = self.client.get("/export?org=UNI-Belgium&sector=Finance")
            self.assertEqual(res_admin.status_code, 200)
            html_admin = res_admin.get_data(as_text=True)
            self.assertIn("Viewing: <strong>UNI-Belgium</strong>", html_admin)
            self.assertIn("UNI-Belgium_Finance_annual_results.csv", html_admin)
            self.assertIn("Download all tables", html_admin)
            self.assertIn("Generate PDF Report", html_admin)

    def test_set_sector_switching(self):
        with self.client:
            self.client.post("/login", data={"user": "system@example.com", "pincode": "0000"}, follow_redirects=True)
            res = self.client.get("/set-sector?sector=Finance", follow_redirects=True)
            self.assertEqual(res.status_code, 200)
            with self.client.session_transaction() as sess:
                self.assertEqual(sess.get("sector"), "Finance")

    def test_pages_render_without_template_errors(self):
        with self.client:
            # Login as regular user
            self.client.post("/login", data={"user": "BE-ACV-Finance", "pincode": "3456"}, follow_redirects=True)
            for path in ("/home", "/annual-results", "/major-agreements", "/major-companies", "/export"):
                res = self.client.get(path)
                self.assertEqual(res.status_code, 200, f"Failed rendering {path} for regular user")

        with self.client:
            # Login as admin
            self.client.post("/login", data={"user": "admin@example.com", "pincode": "1234"}, follow_redirects=True)
            for path in ("/home", "/annual-results", "/major-agreements", "/major-companies", "/export", "/admin"):
                res = self.client.get(path)
                self.assertEqual(res.status_code, 200, f"Failed rendering {path} for admin")

    def test_admin_editing_affiliate_preserves_focus(self):
        with self.client:
            # Login as admin
            self.client.post("/login", data={"user": "admin@example.com", "pincode": "1234"}, follow_redirects=True)

            # View annual results as affiliate
            res = self.client.get("/annual-results?org=UNI-Belgium")
            self.assertEqual(res.status_code, 200)
            self.assertIn("Viewing: <strong>UNI-Belgium</strong>", res.get_data(as_text=True))

            # Save annual results for affiliate
            post_res = self.client.post("/annual-results", data={
                "org": "UNI-Belgium",
                "value_collective_bargaining_coverage_2026": "88",
            }, follow_redirects=False)
            self.assertEqual(post_res.status_code, 302)
            self.assertIn("org=UNI-Belgium", post_res.headers.get("Location", ""))

            # Follow redirect and verify focus is preserved
            follow_res = self.client.get(post_res.headers["Location"])
            self.assertEqual(follow_res.status_code, 200)
            self.assertIn("Viewing: <strong>UNI-Belgium</strong>", follow_res.get_data(as_text=True))

            # Save agreements for affiliate
            ag_post = self.client.post("/major-agreements", data={
                "org": "UNI-Belgium",
                "year": "2026",
                "company_name_0": "Test CLA",
                "workers_affected_0": "500",
            }, follow_redirects=False)
            self.assertEqual(ag_post.status_code, 302)
            self.assertIn("org=UNI-Belgium", ag_post.headers.get("Location", ""))

            # Save companies for affiliate
            co_post = self.client.post("/major-companies", data={
                "org": "UNI-Belgium",
                "year": "2026",
                "company_name_0": "Test Company",
                "workers_0": "1000",
            }, follow_redirects=False)
            self.assertEqual(co_post.status_code, 302)
            self.assertIn("org=UNI-Belgium", co_post.headers.get("Location", ""))

    def test_security_users_csv_cannot_be_downloaded(self):
        with self.client:
            self.client.post("/login", data={"user": "BE-ACV-Finance", "pincode": "3456"}, follow_redirects=True)
            res1 = self.client.get("/export/download/users.csv")
            self.assertIn(res1.status_code, [403, 404])
            res2 = self.client.get("/export/download/../users.csv")
            self.assertIn(res2.status_code, [403, 404, 400])
            res3 = self.client.get("/export/download/..%2Fusers.csv")
            self.assertIn(res3.status_code, [403, 404, 400])
            res4 = self.client.get("/export/download/unicompass.db")
            self.assertIn(res4.status_code, [403, 404])

    def test_admin_download_db_backup(self):
        with self.client:
            # 1. Non-admin user is denied all 4 download routes
            self.client.post("/login", data={"user": "BE-ACV-Finance", "pincode": "3456"}, follow_redirects=True)
            for path in ("/admin/download-db", "/admin/download-csv", "/admin/download-xlsx", "/admin/download-zip"):
                res_user = self.client.get(path)
                self.assertEqual(res_user.status_code, 302)

            # 2. Admin user receives timestamped database files (.db, .csv, .xlsx, .zip)
            self.client.post("/login", data={"user": "admin@example.com", "pincode": "1234"}, follow_redirects=True)
            
            # Check admin page has all 4 download buttons
            admin_page = self.client.get("/admin")
            self.assertEqual(admin_page.status_code, 200)
            admin_html = admin_page.get_data(as_text=True)
            self.assertIn("Download .db", admin_html)
            self.assertIn("Download .csv", admin_html)
            self.assertIn("Download .xlsx", admin_html)
            self.assertIn("Download all tables (.zip)", admin_html)
            self.assertIn("/admin/download-db", admin_html)
            self.assertIn("/admin/download-csv", admin_html)
            self.assertIn("/admin/download-xlsx", admin_html)
            self.assertIn("/admin/download-zip", admin_html)

            # Download the DB copy (.db)
            res_db = self.client.get("/admin/download-db")
            self.assertEqual(res_db.status_code, 200)
            self.assertEqual(res_db.mimetype, "application/x-sqlite3")
            self.assertIn("unicompass_", res_db.headers.get("Content-Disposition", ""))
            self.assertIn(".db", res_db.headers.get("Content-Disposition", ""))

            # Validate SQLite database integrity
            import tempfile
            import sqlite3
            with tempfile.NamedTemporaryFile(suffix=".db", delete=False) as tmp:
                tmp.write(res_db.data)
                tmp_path = tmp.name

            try:
                conn = sqlite3.connect(tmp_path)
                cur = conn.cursor()
                cur.execute("SELECT name FROM sqlite_master WHERE type='table'")
                tables = [r[0] for r in cur.fetchall()]
                conn.close()
                self.assertIn("users", tables)
                self.assertIn("annual_results", tables)
                self.assertIn("agreements", tables)
                self.assertIn("companies", tables)
            finally:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)

            # Download the CSV ZIP (.zip)
            res_csv = self.client.get("/admin/download-csv")
            self.assertEqual(res_csv.status_code, 200)
            self.assertEqual(res_csv.mimetype, "application/zip")
            self.assertIn("unicompass_csv_", res_csv.headers.get("Content-Disposition", ""))
            self.assertIn(".zip", res_csv.headers.get("Content-Disposition", ""))
            
            from zipfile import ZipFile
            from io import BytesIO
            with ZipFile(BytesIO(res_csv.data), "r") as zf:
                names = zf.namelist()
                self.assertIn("annual_results.csv", names)
                self.assertIn("agreements.csv", names)
                self.assertIn("companies.csv", names)
                self.assertIn("users.csv", names)

            # Download the Excel workbook (.xlsx)
            res_xlsx = self.client.get("/admin/download-xlsx")
            self.assertEqual(res_xlsx.status_code, 200)
            self.assertEqual(res_xlsx.mimetype, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.assertIn("unicompass_", res_xlsx.headers.get("Content-Disposition", ""))
            self.assertIn(".xlsx", res_xlsx.headers.get("Content-Disposition", ""))
            
            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(res_xlsx.data))
            self.assertIn("annual_results", wb.sheetnames)
            self.assertIn("agreements", wb.sheetnames)
            self.assertIn("companies", wb.sheetnames)
            self.assertIn("users", wb.sheetnames)

            # Download the all tables ZIP (.zip)
            res_all_zip = self.client.get("/admin/download-zip")
            self.assertEqual(res_all_zip.status_code, 200)
            self.assertEqual(res_all_zip.mimetype, "application/zip")
            self.assertIn("unicompass_all_tables_", res_all_zip.headers.get("Content-Disposition", ""))
            self.assertIn(".zip", res_all_zip.headers.get("Content-Disposition", ""))

            with ZipFile(BytesIO(res_all_zip.data), "r") as zf:
                all_names = zf.namelist()
                self.assertIn("annual_results.csv", all_names)
                self.assertIn("unicompass.xlsx", all_names)
                self.assertTrue(any("UNI-Belgium" in name for name in all_names))

    def test_admin_viewing_affiliate_progress_consistency(self):
        with self.client:
            # Login as admin
            self.client.post("/login", data={"user": "admin@example.com", "pincode": "1234"}, follow_redirects=True)
            
            # Switch to UNI-Turkey and Commerce
            res = self.client.get("/home?org=UNI-Turkey", follow_redirects=True)
            self.assertEqual(res.status_code, 200)
            self.assertIn("Viewing: <strong>UNI-Turkey</strong>", res.get_data(as_text=True))
            
            # Progress table should compute for UNI-Turkey (0 completed), not UNI Europe
            res_turkey_home = self.client.get("/home")
            self.assertEqual(res_turkey_home.status_code, 200)
            self.assertIn("Viewing: <strong>UNI-Turkey</strong>", res_turkey_home.get_data(as_text=True))

    def test_pdf_report_generation_and_export(self):
        # 1. Direct model PDF generation
        from models import generate_pdf_report, save_agreements, save_companies
        
        # Save agreement and company with multi-line comment & other_changes
        save_agreements("UNI-Belgium", 2026, [
            {
                "company_name": "BelgBank SA",
                "level": "Single-employer",
                "date_of_agreement": "2026-06",
                "duration": "2 years",
                "workers_affected": 4200,
                "wage_increase": "4.2%",
                "one_off_lump_sum": "500 €",
                "other_changes": "Training & working time\nDetailed breakdown: 35h week and 5 days training.",
                "comment": "Major breakthrough\nNegotiations lasted 6 months but settled amicably."
            }
        ], sector="Finance")
        
        save_companies("UNI-Belgium", 2026, [
            {
                "company_name": "BelgBank SA",
                "workers": 4500,
                "mnc": "Yes",
                "agreement": "Yes",
                "number_of_unions": 3,
                "union_density": "75-89%",
                "worker_representation": "Yes",
                "ewc_presence": "Yes",
                "comment": "Strong union presence\nWorks council meets bi-monthly."
            }
        ], sector="Finance")

        pdf_bytes = generate_pdf_report(organisation="UNI-Belgium", sector="Finance")
        self.assertTrue(pdf_bytes.startswith(b"%PDF-"))
        self.assertGreater(len(pdf_bytes), 1000)

        # 2. Regular user endpoint access
        with self.client:
            self.client.post("/login", data={"user": "BE-ACV-Finance", "pincode": "3456"}, follow_redirects=True)
            
            # Export page loads and lists files (PDF is excluded from table, accessed via button)
            res_export = self.client.get("/export")
            self.assertEqual(res_export.status_code, 200)
            self.assertIn("Generate PDF Report", res_export.get_data(as_text=True))
            self.assertIn("UNI-Belgium_Finance_annual_results.csv", res_export.get_data(as_text=True))
            self.assertNotIn("UNI-Belgium_Finance_report.pdf", res_export.get_data(as_text=True))

            # Direct PDF generation endpoint
            res_pdf = self.client.get("/export/pdf")
            self.assertEqual(res_pdf.status_code, 200)
            self.assertEqual(res_pdf.mimetype, "application/pdf")
            self.assertIn("UNI-Belgium_Finance_report.pdf", res_pdf.headers.get("Content-Disposition", ""))
            self.assertTrue(res_pdf.data.startswith(b"%PDF-"))
            res_pdf.close()

            # Download pre-generated PDF file
            res_dl = self.client.get("/export/download/UNI-Belgium_Finance_report.pdf")
            self.assertEqual(res_dl.status_code, 200)
            self.assertTrue(res_dl.data.startswith(b"%PDF-"))
            res_dl.close()

            # Cannot download other affiliate's PDF
            res_forbidden = self.client.get("/export/download/UNI-Turkey_Commerce_report.pdf")
            self.assertEqual(res_forbidden.status_code, 403)
            res_forbidden.close()

        # 3. Admin PDF generation endpoint
        with self.client:
            self.client.post("/login", data={"user": "admin@example.com", "pincode": "1234"}, follow_redirects=True)
            res_admin_pdf = self.client.get("/export/pdf?org=UNI-Turkey&sector=Commerce")
            self.assertEqual(res_admin_pdf.status_code, 200)
            self.assertEqual(res_admin_pdf.mimetype, "application/pdf")
            self.assertIn("UNI-Turkey_Commerce_report.pdf", res_admin_pdf.headers.get("Content-Disposition", ""))
            self.assertTrue(res_admin_pdf.data.startswith(b"%PDF-"))
            res_admin_pdf.close()

            # Download all zip should contain PDF reports
            res_zip = self.client.get("/export/download-all")
            self.assertEqual(res_zip.status_code, 200)
            from zipfile import ZipFile
            from io import BytesIO
            with ZipFile(BytesIO(res_zip.data), "r") as zf:
                pdf_entries = [name for name in zf.namelist() if name.endswith(".pdf")]
                self.assertGreater(len(pdf_entries), 0)
            res_zip.close()

    def test_autosave_ajax_and_title_content_separation(self):
        with self.client:
            self.client.post("/login", data={"user": "BE-ACV-Finance", "pincode": "3456"}, follow_redirects=True)

            # 1. AJAX auto-save on Annual results with distinct title and content
            res_ann = self.client.post(
                "/annual-results",
                data={
                    "value_bargaining_climate_2026": "",
                    "qual_bargaining_climate_2026": "rather favourable",
                    "title_outcome_reporting_2026": "Pensions and wage framework",
                    "content_outcome_reporting_2026": "A breakthrough was reached after multi-month tripartite discussions.",
                },
                headers={"X-Requested-With": "XMLHttpRequest"}
            )
            self.assertEqual(res_ann.status_code, 200)
            self.assertEqual(res_ann.json.get("status"), "ok")

            # Verify in GET request that title and content are distinct
            res_get = self.client.get("/annual-results")
            html = res_get.get_data(as_text=True)
            self.assertIn('value="Pensions and wage framework"', html)
            self.assertIn('A breakthrough was reached after multi-month tripartite discussions.', html)

            # 2. AJAX auto-save on Major agreements
            res_agr = self.client.post(
                "/major-agreements",
                data={
                    "year": "2026",
                    "company_name_0": "FinanceCorp SA",
                    "workers_affected_0": "3200",
                    "wage_increase_0": "3.8%",
                    "comment_title_0": "Stable deal",
                    "comment_content_0": "Concluded with high union satisfaction.",
                },
                headers={"X-Requested-With": "XMLHttpRequest"}
            )
            self.assertEqual(res_agr.status_code, 200)
            self.assertEqual(res_agr.json.get("status"), "ok")

            # 3. AJAX auto-save on Major companies
            res_comp = self.client.post(
                "/major-companies",
                data={
                    "year": "2026",
                    "company_name_0": "GlobalBank Ltd",
                    "workers_0": "5000",
                    "comment_title_0": "EWC meeting scheduled",
                    "comment_content_0": "Annual plenary scheduled for November.",
                },
                headers={"X-Requested-With": "XMLHttpRequest"}
            )
            self.assertEqual(res_comp.status_code, 200)
            self.assertEqual(res_comp.json.get("status"), "ok")

    def test_five_strikes_lockout_and_9999_pincode_blocked(self):
        test_user = "test@example.com"
        orig_pin = "5678"
        try:
            # Ensure starting state
            set_user_pincode(test_user, orig_pin)
            clear_login_attempts(test_user)

            # 1. Pincode 9999 is always blocked
            res_9999 = self.client.post("/login", data={"user": test_user, "pincode": "9999"}, follow_redirects=True)
            self.assertIn("Invalid username or pincode", res_9999.get_data(as_text=True))
            clear_login_attempts(test_user)

            # 2. Attempt 4 failed logins with wrong pincode
            for i in range(4):
                res = self.client.post("/login", data={"user": test_user, "pincode": "0001"}, follow_redirects=True)
                self.assertIn("Invalid username or pincode", res.get_data(as_text=True))
                user_info = get_user(test_user)
                self.assertEqual(user_info["pincode"], orig_pin)

            # 3. Fifth failed attempt triggers lockout and sets pincode to 9999
            res_5 = self.client.post("/login", data={"user": test_user, "pincode": "0001"}, follow_redirects=True)
            html_5 = res_5.get_data(as_text=True)
            self.assertIn("contact the administrator for a new pincode", html_5)

            # Verify in users.csv that pincode is now 9999
            user_info = get_user(test_user)
            self.assertEqual(user_info["pincode"], "9999")
            self.assertTrue(is_user_blocked(test_user))

            # 4. Subsequent login attempt (even with original pin or 9999) is blocked with administrator contact message
            res_locked = self.client.post("/login", data={"user": test_user, "pincode": orig_pin}, follow_redirects=True)
            html_locked = res_locked.get_data(as_text=True)
            self.assertIn("contact the administrator for a new pincode", html_locked)

            # 5. Admin resets user pincode
            set_user_pincode(test_user, "4321")
            clear_login_attempts(test_user)
            res_restored = self.client.post("/login", data={"user": test_user, "pincode": "4321"}, follow_redirects=True)
            self.assertIn(f"Welcome, {test_user}!", res_restored.get_data(as_text=True))

        finally:
            # Restore original pin in users.csv
            set_user_pincode(test_user, orig_pin)
            clear_login_attempts(test_user)


if __name__ == "__main__":
    unittest.main()


