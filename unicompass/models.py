"""
Database models for UNI Compass.
Uses SQLite via the sqlite3 standard library — no ORM required.
"""

import csv
import sqlite3
import os
import datetime
import tempfile
from zipfile import ZipFile
from io import BytesIO, StringIO
from contextlib import contextmanager
from config import DATABASE, EXPORT_DIR


@contextmanager
def get_db():
    """Open a connection to the SQLite database, yielding a transaction context and closing it on exit."""
    os.makedirs(os.path.dirname(DATABASE), exist_ok=True)
    conn = sqlite3.connect(DATABASE, timeout=30.0)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout = 30000")
    conn.execute("PRAGMA foreign_keys=ON")
    try:
        with conn:
            yield conn
    finally:
        conn.close()


def init_db():
    """Create all tables if they do not exist."""
    with get_db() as db:
        db.executescript("""
            CREATE TABLE IF NOT EXISTS users (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                user        TEXT    NOT NULL UNIQUE,
                pincode     TEXT    NOT NULL,
                organisation TEXT   NOT NULL,
                is_admin    INTEGER NOT NULL DEFAULT 0,
                country     TEXT    DEFAULT '',
                sector      TEXT    DEFAULT ''
            );

            CREATE TABLE IF NOT EXISTS login_attempts (
                id       INTEGER PRIMARY KEY AUTOINCREMENT,
                user     TEXT    NOT NULL,
                attempted_at TEXT NOT NULL DEFAULT (datetime('now'))
            );

            CREATE TABLE IF NOT EXISTS login_bans (
                user     TEXT    PRIMARY KEY,
                banned_until TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS annual_results (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                organisation TEXT    NOT NULL,
                variable     TEXT    NOT NULL,
                year         INTEGER NOT NULL,
                value        TEXT    DEFAULT '',
                comment      TEXT    DEFAULT '',
                qual_value   TEXT    DEFAULT '',
                row_mode     TEXT    DEFAULT '',
                UNIQUE(organisation, variable, year)
            );

            CREATE TABLE IF NOT EXISTS agreements (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                organisation    TEXT    NOT NULL,
                year            INTEGER NOT NULL,
                row_idx         INTEGER NOT NULL,
                company_name    TEXT    DEFAULT '',
                wage_increase   TEXT    DEFAULT '',
                workers_affected INTEGER DEFAULT 0,
                level           TEXT    DEFAULT '',
                date_of_agreement TEXT   DEFAULT '',
                one_off_lump_sum TEXT    DEFAULT '',
                bargaining_climate TEXT DEFAULT '',
                comment         TEXT    DEFAULT '',
                duration        TEXT    DEFAULT '',
                working_time_change TEXT DEFAULT '',
                other_changes   TEXT    DEFAULT '',
                one_off_bonus   TEXT    DEFAULT ''
            );

            CREATE TABLE IF NOT EXISTS companies (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                organisation TEXT    NOT NULL,
                year         INTEGER NOT NULL,
                row_idx      INTEGER NOT NULL,
                company_name TEXT    DEFAULT '',
                agreement    TEXT    DEFAULT '',
                workers      INTEGER DEFAULT 0,
                number_of_unions INTEGER DEFAULT 0,
                union_members    INTEGER DEFAULT 0,
                ewc_presence  TEXT   DEFAULT '',
                mnc           TEXT   DEFAULT '',
                bargaining_climate TEXT DEFAULT '',
                one_off_lump_sum TEXT   DEFAULT '',
                comment      TEXT    DEFAULT '',
                union_density TEXT   DEFAULT '',
                worker_representation TEXT DEFAULT '',
                one_off_bonus TEXT   DEFAULT ''
            );
        """)
        # Migrate existing tables that may lack new columns
        for table, col in (("agreements", "bargaining_climate"),
                           ("agreements", "duration"),
                           ("agreements", "working_time_change"),
                           ("agreements", "other_changes"),
                           ("agreements", "one_off_lump_sum"),
                           ("agreements", "one_off_bonus"),
                           ("companies", "ewc_presence"),
                           ("companies", "mnc"),
                           ("companies", "bargaining_climate"),
                           ("companies", "union_density"),
                           ("companies", "worker_representation"),
                           ("companies", "one_off_lump_sum"),
                           ("companies", "one_off_bonus"),
                           ("users", "country"),
                           ("users", "sector"),
                           ("annual_results", "qual_value"),
                           ("annual_results", "row_mode"),
                           ("annual_results", "sector"),
                           ("agreements", "sector"),
                           ("companies", "sector")):
            try:
                db.execute(f"ALTER TABLE {table} ADD COLUMN {col} TEXT DEFAULT ''")
            except Exception:
                pass

        # Migrate existing annual results avg_wage_increase variable to avg_pay_increase
        try:
            db.execute("UPDATE annual_results SET variable = 'avg_pay_increase' WHERE variable = 'avg_wage_increase'")
        except Exception:
            pass

        # Migrate existing annual results one_off_bonus variable to one_off_lump_sum
        try:
            db.execute("UPDATE annual_results SET variable = 'one_off_lump_sum' WHERE variable = 'one_off_bonus'")
        except Exception:
            pass        # Migrate annual_results table if it lacks sector in UNIQUE constraint
        try:
            # Check table definition
            tbl_sql = db.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='annual_results'").fetchone()
            if tbl_sql and "UNIQUE(organisation, variable, year)" in tbl_sql[0]:
                db.executescript("""
                    CREATE TABLE annual_results_migrated (
                        id           INTEGER PRIMARY KEY AUTOINCREMENT,
                        organisation TEXT    NOT NULL,
                        sector       TEXT    DEFAULT '',
                        variable     TEXT    NOT NULL,
                        year         INTEGER NOT NULL,
                        value        TEXT    DEFAULT '',
                        comment      TEXT    DEFAULT '',
                        qual_value   TEXT    DEFAULT '',
                        row_mode     TEXT    DEFAULT '',
                        UNIQUE(organisation, sector, variable, year)
                    );
                    INSERT OR IGNORE INTO annual_results_migrated (id, organisation, sector, variable, year, value, comment, qual_value, row_mode)
                    SELECT id, organisation, COALESCE(sector, ''), variable, year, value, comment, COALESCE(qual_value, ''), COALESCE(row_mode, '') FROM annual_results;
                    DROP TABLE annual_results;
                    ALTER TABLE annual_results_migrated RENAME TO annual_results;
                """)
        except Exception:
            pass

        # Migrate old agreements and companies one_off_bonus values to one_off_lump_sum
        try:
            db.execute("UPDATE agreements SET one_off_lump_sum = one_off_bonus WHERE (one_off_lump_sum = '' OR one_off_lump_sum IS NULL) AND one_off_bonus != '' AND one_off_bonus IS NOT NULL")
            db.execute("UPDATE companies SET one_off_lump_sum = one_off_bonus WHERE (one_off_lump_sum = '' OR one_off_lump_sum IS NULL) AND one_off_bonus != '' AND one_off_bonus IS NOT NULL")
        except Exception:
            pass

        # Migrate transient login_bans and login_attempts tables to use user
        try:
            tbl_sql = db.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='login_bans'").fetchone()
            if tbl_sql and "email" in tbl_sql[0]:
                db.execute("DROP TABLE IF EXISTS login_bans")
                db.execute("CREATE TABLE login_bans (user TEXT PRIMARY KEY, banned_until TEXT NOT NULL)")
        except Exception:
            pass

        try:
            tbl_sql = db.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='login_attempts'").fetchone()
            if tbl_sql and "email" in tbl_sql[0]:
                db.execute("DROP TABLE IF EXISTS login_attempts")
                db.execute("CREATE TABLE login_attempts (id INTEGER PRIMARY KEY AUTOINCREMENT, user TEXT NOT NULL, attempted_at TEXT NOT NULL DEFAULT (datetime('now')))")
        except Exception:
            pass


def check_login_rate_limit(user):
    """Return True if the user is allowed to attempt login, False if rate-limited."""
    user = user.strip().lower()
    with get_db() as db:
        ban = db.execute("SELECT banned_until FROM login_bans WHERE user = ?", (user,)).fetchone()
        if ban:
            now_iso = datetime.datetime.now(datetime.timezone.utc).isoformat()
            if now_iso < ban["banned_until"]:
                return False
            db.execute("DELETE FROM login_bans WHERE user = ?", (user,))
        return True


def record_failed_login(user):
    """Record a failed login attempt; ban for 15 minutes after 5 failures within 15 minutes."""
    user = user.strip().lower()
    now_utc = datetime.datetime.now(datetime.timezone.utc)
    with get_db() as db:
        db.execute("INSERT INTO login_attempts (user, attempted_at) VALUES (?, ?)", (user, now_utc.isoformat()))
        cutoff = (now_utc - datetime.timedelta(minutes=15)).isoformat()
        count = db.execute(
            "SELECT COUNT(*) FROM login_attempts WHERE user = ? AND attempted_at >= ?",
            (user, cutoff)
        ).fetchone()[0]
        if count >= 5:
            banned_until = (now_utc + datetime.timedelta(minutes=15)).isoformat()
            db.execute(
                "INSERT OR REPLACE INTO login_bans (user, banned_until) VALUES (?, ?)",
                (user, banned_until)
            )


def clear_login_attempts(user):
    """Clear failed login attempts for a user after a successful login."""
    user = user.strip().lower()
    with get_db() as db:
        db.execute("DELETE FROM login_attempts WHERE user = ?", (user,))
        db.execute("DELETE FROM login_bans WHERE user = ?", (user,))


def get_annual_results(organisation, years=None, sector=None):
    """Return a dict: (variable, year) -> {'value': ..., 'comment': ..., 'qual_value': ..., 'row_mode': ...}"""
    if isinstance(years, int):
        years = [years]

    qual_words = {
        "100%", "very high", "high", "average", "normal", "moderate", "low", "very low", "0%", "none",
        "80-100%", "60-79%", "40-59%", "20-39%", "0-19%",
        "strong increase", "moderate increase", "minimal increase", "no increase", "decrease",
        "minor increase", "stable", "minor decrease", "strong decrease",
        "many agreements / many employers", "many agreements / few employers",
        "few agreements / many employers", "few agreements / few employers", "no agreements",
        "many agreements/many employers", "many agreements/few employers",
        "few agreements/many employers", "few agreements/few employers",
        "very many agreements", "many agreements", "few agreements", "very few agreements",
        "very common", "common", "rare", "very rare",
        "strong inflation (+3%)", "minor inflation (+1/+3%)", "stable prices (+1/-1%)",
        "minor deflation (-1/-3%)", "strong deflation (-3%)",
        "strong inflation", "minor inflation", "stable prices", "minor deflation", "strong deflation",
        "strong growth", "moderate growth", "no growth", "moderate decline", "strong decline",
        "very unfavourable", "rather unfavourable", "neutral", "rather favourable", "very favourable",
        "very difficult", "difficult", "favourable"
    }

    with get_db() as db:
        if years:
            placeholders = ",".join("?" for _ in years)
            if sector and sector.lower() != "all":
                rows = db.execute(
                    f"SELECT variable, year, value, comment, qual_value, row_mode FROM annual_results "
                    f"WHERE organisation = ? AND sector = ? AND year IN ({placeholders})",
                    (organisation, sector, *years)
                ).fetchall()
            else:
                rows = db.execute(
                    f"SELECT variable, year, value, comment, qual_value, row_mode FROM annual_results "
                    f"WHERE organisation = ? AND year IN ({placeholders})",
                    (organisation, *years)
                ).fetchall()
        else:
            if sector and sector.lower() != "all":
                rows = db.execute(
                    "SELECT variable, year, value, comment, qual_value, row_mode FROM annual_results "
                    "WHERE organisation = ? AND sector = ?",
                    (organisation, sector)
                ).fetchall()
            else:
                rows = db.execute(
                    "SELECT variable, year, value, comment, qual_value, row_mode FROM annual_results "
                    "WHERE organisation = ?",
                    (organisation,)
                ).fetchall()
    result = {}
    for r in rows:
        val = r["value"] or ""
        qual = r["qual_value"] or ""
        comment = r["comment"] or ""
        row_mode = r["row_mode"] or ""
        
        # Backward compatibility check:
        if val in qual_words and not qual:
            qual = val
            val = ""
            
        result[(r["variable"], r["year"])] = {
            "value": val,
            "qual_value": qual,
            "row_mode": row_mode,
            "comment": comment
        }
    return result


def save_annual_results(organisation, data, sector=""):
    """data is list of dicts: [{'variable': ..., 'year': ..., 'value': ..., 'qual_value': ..., 'row_mode': ..., 'comment': ...}]"""
    sec = sector if (sector and sector.lower() != "all") else ""
    with get_db() as db:
        for d in data:
            item_sec = sec or d.get("sector", "")
            db.execute(
                "DELETE FROM annual_results WHERE organisation = ? AND sector = ? AND variable = ? AND year = ?",
                (organisation, item_sec, d["variable"], d["year"])
            )
            db.execute(
                "INSERT INTO annual_results (organisation, sector, variable, year, value, comment, qual_value, row_mode) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (organisation, item_sec, d["variable"], d["year"], d.get("value", ""), d.get("comment", ""),
                 d.get("qual_value", ""), d.get("row_mode", ""))
            )


def get_agreements(organisation, year, sector=None):
    """Return rows sorted by workers_affected descending."""
    with get_db() as db:
        if sector and sector.lower() != "all":
            rows = db.execute(
                "SELECT * FROM agreements WHERE organisation = ? AND sector = ? AND year = ? "
                "ORDER BY workers_affected DESC, row_idx",
                (organisation, sector, year)
            ).fetchall()
        else:
            rows = db.execute(
                "SELECT * FROM agreements WHERE organisation = ? AND year = ? "
                "ORDER BY workers_affected DESC, row_idx",
                (organisation, year)
            ).fetchall()
    return [dict(r) for r in rows]


def save_agreements(organisation, year, rows, sector=""):
    """Replace all agreement rows for this org+sector+year."""
    sec = sector if (sector and sector.lower() != "all") else ""
    with get_db() as db:
        db.execute("DELETE FROM agreements WHERE organisation = ? AND sector = ? AND year = ?", (organisation, sec, year))
        for i, r in enumerate(rows):
            db.execute(
                "INSERT INTO agreements (organisation, sector, year, row_idx, company_name, wage_increase, workers_affected, level, date_of_agreement, one_off_lump_sum, bargaining_climate, comment, duration, working_time_change, other_changes) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (organisation, sec, year, i, r.get("company_name", ""), r.get("wage_increase", ""),
                 r.get("workers_affected", 0), r.get("level", ""), r.get("date_of_agreement", ""),
                 r.get("one_off_lump_sum", ""), r.get("bargaining_climate", ""), r.get("comment", ""),
                 r.get("duration", ""), r.get("working_time_change", ""), r.get("other_changes", ""))
            )


def get_companies(organisation, year, sector=None):
    """Return rows sorted by workers descending."""
    with get_db() as db:
        if sector and sector.lower() != "all":
            rows = db.execute(
                "SELECT * FROM companies WHERE organisation = ? AND sector = ? AND year = ? "
                "ORDER BY workers DESC, row_idx",
                (organisation, sector, year)
            ).fetchall()
        else:
            rows = db.execute(
                "SELECT * FROM companies WHERE organisation = ? AND year = ? "
                "ORDER BY workers DESC, row_idx",
                (organisation, year)
            ).fetchall()
    return [dict(r) for r in rows]


def save_companies(organisation, year, rows, sector=""):
    """Replace all company rows for this org+sector+year."""
    sec = sector if (sector and sector.lower() != "all") else ""
    with get_db() as db:
        db.execute("DELETE FROM companies WHERE organisation = ? AND sector = ? AND year = ?", (organisation, sec, year))
        for i, r in enumerate(rows):
            db.execute(
                "INSERT INTO companies (organisation, sector, year, row_idx, company_name, agreement, workers, number_of_unions, union_members, ewc_presence, mnc, bargaining_climate, one_off_lump_sum, comment, union_density, worker_representation) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (organisation, sec, year, i, r.get("company_name", ""), r.get("agreement", ""),
                 r.get("workers", 0), r.get("number_of_unions", 0), r.get("union_members", 0),
                 r.get("ewc_presence", ""), r.get("mnc", ""), r.get("bargaining_climate", ""),
                 r.get("one_off_lump_sum", ""), r.get("comment", ""), r.get("union_density", ""), r.get("worker_representation", ""))
            )


def get_all_organisations():
    """Return sorted list of distinct organisation names from results tables."""
    orgs = set()
    with get_db() as db:
        for table in ("annual_results", "agreements", "companies"):
            rows = db.execute(f"SELECT DISTINCT organisation FROM {table}").fetchall()
            orgs.update(r["organisation"] for r in rows)
    return sorted(orgs)


def compute_progress(org, years, sector=None):
    """Return dict with completion counts for an organisation."""
    annual_total = 0
    annual_filled = 0
    for y in years:
        data = get_annual_results(org, [y], sector=sector)
        for var_key in ("bargaining_climate", "collective_bargaining_coverage", "avg_pay_increase",
                        "one_off_lump_sum", "multi_employer_agreements", "single_employer_agreements",
                        "avg_monthly_wage", "inflation_rate", "sectoral_productivity", "outcome_reporting"):
            annual_total += 1
            cell = data.get((var_key, y), {})
            if cell.get("value", "") or cell.get("qual_value", "") or cell.get("comment", ""):
                annual_filled += 1

    agreements_filled = 0
    for y in years:
        rows = get_agreements(org, y, sector=sector)
        for r in rows:
            if any(r.get(k) for k in ("company_name", "wage_increase", "workers_affected", "one_off_lump_sum", "other_changes", "comment")):
                agreements_filled += 1

    companies_filled = 0
    for y in years:
        rows = get_companies(org, y, sector=sector)
        for r in rows:
            if any(r.get(k) for k in ("company_name", "workers", "agreement", "union_density", "comment")):
                companies_filled += 1

    return {
        "annual": {"filled": annual_filled, "total": annual_total},
        "agreements": {"filled": agreements_filled, "total": agreements_filled},
        "companies": {"filled": companies_filled, "total": companies_filled},
    }


def _export_xlsx(organisation=None, sector=None):
    """Write tables to a single XLSX workbook. If organisation and sector are given, filter by both."""
    from openpyxl import Workbook
    wb = Workbook()
    default = wb.active
    sheet_count = 0
    with get_db() as db:
        for table in ("annual_results", "agreements", "companies"):
            if organisation and sector and sector.lower() != "all":
                rows = db.execute(
                    f"SELECT * FROM {table} WHERE organisation = ? AND sector = ? ORDER BY organisation, year",
                    (organisation, sector)
                ).fetchall()
            elif organisation:
                rows = db.execute(
                    f"SELECT * FROM {table} WHERE organisation = ? ORDER BY organisation, year",
                    (organisation,)
                ).fetchall()
            else:
                rows = db.execute(f"SELECT * FROM {table} ORDER BY organisation, sector, year").fetchall()
            if not rows:
                continue
            ws = wb.create_sheet(title=table)
            sheet_count += 1
            headers = list(rows[0].keys())
            ws.append(headers)
            for r in rows:
                ws.append([r[h] for h in headers])
    if sheet_count == 0:
        return b""
    wb.remove(default)
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def org_file_prefix(org):
    """Safe filename prefix for an organisation (spaces → underscores)."""
    return org.replace(" ", "_") if org else ""


def org_sector_file_prefix(org, sector=""):
    """Safe filename prefix for an organisation and sector combination."""
    safe_org = (org or "").replace(" ", "_")
    if sector and sector.lower() != "all":
        safe_sec = sector.replace(" ", "_")
        return f"{safe_org}_{safe_sec}"
    return safe_org


def export_all():
    """Write full database files and per-affiliate/sector CSVs and XLSX to EXPORT_DIR."""
    os.makedirs(EXPORT_DIR, exist_ok=True)
    pairs_set = set()
    orgs_set = set()

    with get_db() as db:
        for table in ("annual_results", "agreements", "companies"):
            rows = [dict(r) for r in db.execute(f"SELECT * FROM {table} ORDER BY organisation, sector, year").fetchall()]
            if rows:
                # 1. Full database CSV (all affiliates and sectors)
                all_path = os.path.join(EXPORT_DIR, f"{table}.csv")
                with open(all_path, "w", newline="") as f:
                    writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
                    writer.writeheader()
                    writer.writerows(rows)
                
                # 2. Collect unique (org, sector) pairs
                for r in rows:
                    o = r.get("organisation", "").strip()
                    s = r.get("sector", "").strip()
                    if o:
                        orgs_set.add(o)
                        pairs_set.add((o, s))

                # 3. Per-(org, sector) and per-org CSV files
                for org, sec in pairs_set:
                    prefix = org_sector_file_prefix(org, sec)
                    filtered_rows = [
                        r for r in rows
                        if r.get("organisation") == org and (not sec or r.get("sector", "") == sec or sec.lower() == "all")
                    ]
                    if filtered_rows:
                        pair_path = os.path.join(EXPORT_DIR, f"{prefix}_{table}.csv")
                        with open(pair_path, "w", newline="") as f:
                            writer = csv.DictWriter(f, fieldnames=list(filtered_rows[0].keys()))
                            writer.writeheader()
                            writer.writerows(filtered_rows)

    # 4. Full database XLSX (all affiliates and sectors)
    full_xlsx = _export_xlsx()
    if full_xlsx:
        with open(os.path.join(EXPORT_DIR, "unicompass.xlsx"), "wb") as f:
            f.write(full_xlsx)

    # 5. Per-(org, sector) and per-org XLSX workbooks
    for org, sec in pairs_set:
        prefix = org_sector_file_prefix(org, sec)
        pair_xlsx = _export_xlsx(organisation=org, sector=sec)
        if pair_xlsx:
            with open(os.path.join(EXPORT_DIR, f"{prefix}_unicompass.xlsx"), "wb") as f:
                f.write(pair_xlsx)

    for org in orgs_set:
        safe = org_file_prefix(org)
        org_xlsx = _export_xlsx(organisation=org)
        if org_xlsx:
            with open(os.path.join(EXPORT_DIR, f"{safe}_unicompass.xlsx"), "wb") as f:
                f.write(org_xlsx)

    return sorted(orgs_set)


def generate_export_zip(org_filter=None, sector_filter=None, is_admin=False):
    """Create a zip archive of export files for a specific organisation and sector."""
    export_all()
    buf = BytesIO()
    with ZipFile(buf, "w") as zf:
        if is_admin and not org_filter:
            # Full database dump for administrator
            for fname in ("unicompass.xlsx", "unicompass_report.pdf", "annual_results.csv", "agreements.csv", "companies.csv"):
                fpath = os.path.join(EXPORT_DIR, fname)
                if os.path.isfile(fpath) and fname not in zf.namelist():
                    zf.write(fpath, arcname=fname)
        else:
            prefix = org_sector_file_prefix(org_filter, sector_filter) + "_" if org_filter else None
            for fname in os.listdir(EXPORT_DIR):
                if not (fname.endswith(".csv") or fname.endswith(".xlsx") or fname.endswith(".pdf")):
                    continue
                if prefix and not fname.startswith(prefix):
                    continue
                fpath = os.path.join(EXPORT_DIR, fname)
                if os.path.isfile(fpath) and fname not in zf.namelist():
                    zf.write(fpath, arcname=fname)
    return buf.getvalue()


def generate_all_exports_zip():
    """Create a zip archive of all export files (CSVs, XLSX, and PDFs) for all organisations and sectors in EXPORT_DIR."""
    export_all()
    buf = BytesIO()
    with ZipFile(buf, "w") as zf:
        for fname in sorted(os.listdir(EXPORT_DIR)):
            if fname.endswith(".csv") or fname.endswith(".xlsx") or fname.endswith(".pdf"):
                fpath = os.path.join(EXPORT_DIR, fname)
                if os.path.isfile(fpath) and fname not in zf.namelist():
                    zf.write(fpath, arcname=fname)
    return buf.getvalue()


ANNUAL_VARIABLES_EXPORT = [
    ("bargaining_climate", "Bargaining climate"),
    ("collective_bargaining_coverage", "Collective bargaining coverage (%)"),
    ("avg_pay_increase", "Collectively agreed pay increase (%)"),
    ("one_off_lump_sum", "Additional lump sum (€)"),
    ("multi_employer_agreements", "Number of multi-employer agreements"),
    ("single_employer_agreements", "Number of single-employer agreements"),
    ("avg_monthly_wage", "Average full-time monthly wage (€)"),
    ("inflation_rate", "National inflation rate (%)"),
    ("sectoral_productivity", "Sectoral productivity evolution (%)"),
    ("outcome_reporting", "Qualitative reporting"),
]


def _extract_title(text):
    """Return only the title (first line) of a multi-line comment or other aspects field."""
    if not text:
        return ""
    lines = str(text).strip().split("\n")
    return lines[0].strip() if lines else ""


def generate_pdf_report(organisation=None, sector=None, is_admin=False):
    """Generate a styled landscape A4 PDF report matching the Example layout with title-limited comments."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24
    )
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'RepTitle', parent=styles['Heading1'],
        fontName='Helvetica-Bold', fontSize=15, leading=18,
        textColor=colors.HexColor('#1a3a5c'), spaceAfter=2
    )
    sub_style = ParagraphStyle(
        'RepSub', parent=styles['Normal'],
        fontName='Helvetica', fontSize=8.5, leading=11,
        textColor=colors.HexColor('#555555'), spaceAfter=8
    )
    h2_style = ParagraphStyle(
        'RepH2', parent=styles['Heading2'],
        fontName='Helvetica-Bold', fontSize=10.5, leading=13,
        textColor=colors.HexColor('#1a3a5c'), spaceBefore=8, spaceAfter=4
    )
    cell_style = ParagraphStyle(
        'RepCell', parent=styles['Normal'],
        fontName='Helvetica', fontSize=7.5, leading=9.5,
        alignment=TA_LEFT, textColor=colors.HexColor('#222222')
    )
    cell_bold = ParagraphStyle(
        'RepCellBold', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=7.5, leading=9.5,
        alignment=TA_LEFT, textColor=colors.HexColor('#1a3a5c')
    )
    cell_center = ParagraphStyle(
        'RepCellCenter', parent=styles['Normal'],
        fontName='Helvetica', fontSize=7.5, leading=9.5,
        alignment=TA_CENTER, textColor=colors.HexColor('#222222')
    )
    th_style = ParagraphStyle(
        'RepTh', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=8, leading=10,
        alignment=TA_CENTER, textColor=colors.white
    )
    th_left = ParagraphStyle(
        'RepThLeft', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=8, leading=10,
        alignment=TA_LEFT, textColor=colors.white
    )

    story = []
    
    org_label = organisation or ('All Affiliates' if is_admin else 'General')
    sec_label = sector or 'All Sectors'
    now_str = datetime.datetime.now().strftime('%Y-%m-%d')
    
    story.append(Paragraph('UNI Compass — Collective Bargaining Report', title_style))
    story.append(Paragraph(f'Organisation: <b>{org_label}</b> &nbsp;|&nbsp; Sector: <b>{sec_label}</b> &nbsp;|&nbsp; Generated: <b>{now_str}</b>', sub_style))
    story.append(HRFlowable(width='100%', thickness=1, color=colors.HexColor('#1a3a5c'), spaceBefore=0, spaceAfter=6))

    years = list(range(2023, 2028))
    
    # --- 1. Annual Results ---
    story.append(Paragraph('1. Annual Results', h2_style))
    saved_annual = get_annual_results(organisation, years=years, sector=sector) if organisation else {}
    
    ann_headers = [Paragraph('Variable', th_left)] + [Paragraph(str(y), th_style) for y in years]
    ann_data = [ann_headers]
    for vkey, vname in ANNUAL_VARIABLES_EXPORT:
        row = [Paragraph(vname, cell_bold)]
        for y in years:
            cdata = saved_annual.get((vkey, y), {})
            val = cdata.get('value') or cdata.get('qual_value') or ''
            cmt = _extract_title(cdata.get('comment', ''))
            display_text = val
            if cmt and not val:
                display_text = cmt
            row.append(Paragraph(display_text or '—', cell_center))
        ann_data.append(row)
    
    t_ann = Table(ann_data, colWidths=[240] + [110]*len(years))
    t_ann.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3a5c')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dddddd')),
        ('TOPPADDING', (0, 0), (-1, -1), 2.5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2.5),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
    ]))
    story.append(t_ann)
    story.append(Spacer(1, 10))

    # --- 2. Major Agreements ---
    ag_years_with_data = []
    if organisation:
        for y in sorted(years, reverse=True):
            rows = get_agreements(organisation, y, sector=sector)
            valid = [r for r in rows if any(r.get(k) for k in ('company_name', 'wage_increase', 'workers_affected', 'one_off_lump_sum', 'other_changes', 'comment'))]
            if valid:
                ag_years_with_data.append((y, valid))

    if not ag_years_with_data:
        ag_years_with_data = [(years[-1], [])]

    for y, ag_rows in ag_years_with_data:
        story.append(Paragraph(f'2. Major Agreements ({y})', h2_style))
        ag_headers = [
            Paragraph('#', th_style),
            Paragraph('Reference / Company', th_left),
            Paragraph('Level', th_left),
            Paragraph('Date', th_style),
            Paragraph('Duration', th_style),
            Paragraph('Workers', th_style),
            Paragraph('Pay increase', th_style),
            Paragraph('Lump sum', th_style),
            Paragraph('Other aspects', th_left),
            Paragraph('Comments', th_left),
        ]
        ag_table_data = [ag_headers]
        if ag_rows:
            for idx, r in enumerate(ag_rows, 1):
                ag_table_data.append([
                    Paragraph(str(idx), cell_center),
                    Paragraph(r.get('company_name', '') or '—', cell_bold),
                    Paragraph(r.get('level', '') or '—', cell_style),
                    Paragraph(r.get('date_of_agreement', '') or '—', cell_center),
                    Paragraph(r.get('duration', '') or '—', cell_center),
                    Paragraph(str(r.get('workers_affected', '')) if r.get('workers_affected') else '—', cell_center),
                    Paragraph(r.get('wage_increase', '') or '—', cell_center),
                    Paragraph(r.get('one_off_lump_sum', '') or '—', cell_center),
                    Paragraph(_extract_title(r.get('other_changes', '')) or '—', cell_style),
                    Paragraph(_extract_title(r.get('comment', '')) or '—', cell_style),
                ])
        else:
            ag_table_data.append([Paragraph('—', cell_center)] + [Paragraph('No recorded agreements', cell_style)] + [Paragraph('—', cell_center)]*8)

        t_ag = Table(ag_table_data, colWidths=[20, 115, 65, 55, 55, 55, 65, 60, 150, 150])
        t_ag.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3a5c')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dddddd')),
            ('TOPPADDING', (0, 0), (-1, -1), 2.5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2.5),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ]))
        story.append(t_ag)
        story.append(Spacer(1, 10))

    # --- 3. Major Companies ---
    co_years_with_data = []
    if organisation:
        for y in sorted(years, reverse=True):
            rows = get_companies(organisation, y, sector=sector)
            valid = [r for r in rows if any(r.get(k) for k in ('company_name', 'workers', 'agreement', 'union_density', 'comment'))]
            if valid:
                co_years_with_data.append((y, valid))

    if not co_years_with_data:
        co_years_with_data = [(years[-1], [])]

    for y, co_rows in co_years_with_data:
        story.append(Paragraph(f'3. Major Companies ({y})', h2_style))
        co_headers = [
            Paragraph('#', th_style),
            Paragraph('Company name', th_left),
            Paragraph('Workers', th_style),
            Paragraph('MNC', th_style),
            Paragraph('CLA', th_style),
            Paragraph('Unions', th_style),
            Paragraph('Density', th_style),
            Paragraph('Representation', th_style),
            Paragraph('EWC', th_style),
            Paragraph('Comments', th_left),
        ]
        co_table_data = [co_headers]
        if co_rows:
            for idx, r in enumerate(co_rows, 1):
                co_table_data.append([
                    Paragraph(str(idx), cell_center),
                    Paragraph(r.get('company_name', '') or '—', cell_bold),
                    Paragraph(str(r.get('workers', '')) if r.get('workers') else '—', cell_center),
                    Paragraph(r.get('mnc', '') or '—', cell_center),
                    Paragraph(r.get('agreement', '') or '—', cell_center),
                    Paragraph(str(r.get('number_of_unions', '')) if r.get('number_of_unions') else '—', cell_center),
                    Paragraph(r.get('union_density', '') or '—', cell_center),
                    Paragraph(r.get('worker_representation', '') or '—', cell_center),
                    Paragraph(r.get('ewc_presence', '') or '—', cell_center),
                    Paragraph(_extract_title(r.get('comment', '')) or '—', cell_style),
                ])
        else:
            co_table_data.append([Paragraph('—', cell_center)] + [Paragraph('No recorded companies', cell_style)] + [Paragraph('—', cell_center)]*8)

        t_co = Table(co_table_data, colWidths=[20, 140, 55, 45, 65, 45, 65, 75, 45, 235])
        t_co.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3a5c')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dddddd')),
            ('TOPPADDING', (0, 0), (-1, -1), 2.5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2.5),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ]))
        story.append(t_co)
        story.append(Spacer(1, 10))

    doc.build(story)
    return buf.getvalue()


def backup_database():
    """Create a consistent copy of the SQLite database using SQLite's backup API and return its bytes."""
    with tempfile.NamedTemporaryFile(suffix=".db", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        dest = sqlite3.connect(tmp_path)
        with get_db() as src:
            src.backup(dest)
        dest.close()
        with open(tmp_path, "rb") as f:
            data = f.read()
        return data
    finally:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass


def backup_database_csv_zip():
    """Export all database tables into individual CSV files and package them into a zip archive."""
    buf = BytesIO()
    with get_db() as db:
        tables = [r[0] for r in db.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' AND name NOT IN ('login_attempts', 'login_bans')").fetchall()]
        ordered_tables = []
        for pref in ("annual_results", "agreements", "companies", "users"):
            if pref in tables:
                ordered_tables.append(pref)
        for t in tables:
            if t not in ordered_tables:
                ordered_tables.append(t)

        with ZipFile(buf, "w") as zf:
            for table in ordered_tables:
                rows = [dict(r) for r in db.execute(f"SELECT * FROM {table}").fetchall()]
                csv_buf = StringIO()
                if rows:
                    writer = csv.DictWriter(csv_buf, fieldnames=list(rows[0].keys()))
                    writer.writeheader()
                    writer.writerows(rows)
                else:
                    cols = [c[1] for c in db.execute(f"PRAGMA table_info({table})").fetchall()]
                    writer = csv.DictWriter(csv_buf, fieldnames=cols)
                    writer.writeheader()
                zf.writestr(f"{table}.csv", csv_buf.getvalue())
    return buf.getvalue()


def backup_database_xlsx():
    """Export all database tables into an Excel workbook (.xlsx) with a sheet per table."""
    from openpyxl import Workbook
    wb = Workbook()
    default = wb.active
    sheet_count = 0
    with get_db() as db:
        tables = [r[0] for r in db.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' AND name NOT IN ('login_attempts', 'login_bans')").fetchall()]
        ordered_tables = []
        for pref in ("annual_results", "agreements", "companies", "users"):
            if pref in tables:
                ordered_tables.append(pref)
        for t in tables:
            if t not in ordered_tables:
                ordered_tables.append(t)

        for table in ordered_tables:
            rows = [dict(r) for r in db.execute(f"SELECT * FROM {table}").fetchall()]
            ws = wb.create_sheet(title=table)
            sheet_count += 1
            if rows:
                headers = list(rows[0].keys())
                ws.append(headers)
                for r in rows:
                    ws.append([r[h] for h in headers])
            else:
                cols = [c[1] for c in db.execute(f"PRAGMA table_info({table})").fetchall()]
                ws.append(cols)
    if sheet_count > 0 and default in wb.worksheets:
        wb.remove(default)
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


